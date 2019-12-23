from openpyxl import load_workbook
from support_files.Course import Course
import xlrd
import numpy as np
import os

PPF_COLS = {
    'Course': 'G',
    'Grade': 'I',
    'Credits': 'K',
    'Totals': 'M'
}


# Create column hopping function
def colhop(curr, steps=1):

    curr_col = curr[0]
    curr_row = curr[1:]

    letters = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
    loc = 0
    while letters[loc] != curr_col:
        loc += 1
        if loc == len(letters):
            print('Error with column hopping')

    final_pos = loc + steps
    if final_pos < 0 or final_pos >= len(letters):
        print('Error with column hopping')
    else:
        return letters[final_pos] + str(curr_row)


# Create row hopping function
def rowhop(curr, steps=1):
    curr_col = curr[0]
    curr_row = curr[1:]

    dest = int(curr_row) + steps

    if dest <= 0:
        print('Error with row hopping')
    else:
        return curr_col + str(dest)


def exceeds(a, threshold='C-'):

    grades = {
        'A+': 4.3,
        'A': 4.0,
        'A-': 3.7,
        'B+': 3.3,
        'B': 3.0,
        'B-': 2.7,
        'C+': 2.3,
        'C': 2.0,
        'C-': 1.7,
        'D+': 1.3,
        'D': 1.0,
        'D-': 0.7,
        'F': 0.0
    }

    assert grades.keys().__contains__(a), "Cannot compare a non-letter grade (%s) to the letter grade threshold (%s)" % (a, threshold)
    if grades[a] >= grades[threshold]:
        return True
    else:
        return False


def designate_columns(transcript):
    pos = 'A1'
    curr = transcript[pos].value
    cols = {}

    while curr is not None:
        if curr == 'Academic Term Ldescr' or curr == 'Academic Term Sdescr':
            cols['semester'] = pos[0]
        elif curr == 'Effdt Primary Name' or curr == 'Effdt Preferred Name':
            cols['student name'] = pos[0]
        elif curr == 'Netid':
            cols['netid'] = pos[0]
        elif curr == 'Employee Id':
            cols['student id'] = pos[0]
        elif curr == 'Advisor':
            cols['advisor'] = pos[0]
        elif curr == 'Exp Grad Term Ldescr' or curr == 'Exp Grad Term Sdescr':
            cols['grad'] = pos[0]
        elif curr == 'Class Descr':
            cols['desc'] = pos[0]
        elif curr == 'Subject':
            cols['dept'] = pos[0]
        elif curr == 'Catalog Nbr':
            cols['num'] = pos[0]
        elif curr == 'Official Grade':
            cols['grade'] = pos[0]
        elif curr == 'Unt Taken':
            cols['creds'] = pos[0]
        elif curr == 'AP Test Component Ldescr':
            cols['ap desc'] = pos[0]
        elif curr == 'AP Test Sdescr':
            cols['ap y/n'] = pos[0]
        elif curr == 'Score':
            cols['ap score'] = pos[0]
        elif curr == 'TR School Descr':
            cols['transfer'] = pos[0]

        pos = colhop(pos)
        curr = transcript[pos].value

    return cols


def open_transcript_file(filename):

    wb = load_workbook(filename=filename, data_only=True)
    first_transcript = None
    num_transcripts = 0
    for i in range(len(wb.sheetnames)):
        wb.active = i
        f = wb.active
        if f['A1'].value == 'Effdt Preferred Name':
            if first_transcript is None:
                first_transcript = i
            num_transcripts += 1

    return wb, first_transcript, num_transcripts


def open_excel_file(filename):

    wb = load_workbook(filename=filename, data_only=True)
    file = wb.active

    return filename, wb, file


def read_class_from_ppf(ppf, row):
    if isinstance(row, int):
        row = str(row)

    course_loc = PPF_COLS['Course'] + row
    course = ppf[course_loc].value

    grade_loc = PPF_COLS['Grade'] + row
    grade = ppf[grade_loc].value

    creds_loc = PPF_COLS['Credits'] + row
    creds = ppf[creds_loc].value

    c = Course(course, grade, creds)

    if grade is None:
        print(course)
        print(ppf)
        print(row)
    if "AP" in grade:
        c.ap = True

    return c


def determine_ap_course_equiv(ap_desc, score, course_desc):
    # AP Biology
    if ap_desc == "Biology" and score == "4":
        equiv = Course(num="APBIO4", grade=score, creds=4, desc=course_desc, ap=True, ap_ppf_desc="BIO AP")
    elif ap_desc == "Biology" and score == "5":
        equiv = [Course(num="APBIO5-1", grade=score, creds=3, desc=course_desc, ap=True, ap_ppf_desc="BIO AP"),
                 Course(num="APBIO5-2", grade=score, creds=3, desc=course_desc, ap=True, ap_ppf_desc="BIO AP"),
                 Course(num="APBIOLAB", grade=score, creds=2, desc=course_desc, ap=True, ap_ppf_desc="BIO AP")]
    # AP Chemistry
    elif ap_desc == "Chemistry" and score == "5":
        equiv = Course(num="APCHEM", grade=score, creds=4, desc=course_desc, ap=True, ap_ppf_desc="CHEM AP")
    # AP Computer Science
    elif ap_desc == "Computer Science A" and score == "5":
        equiv = Course(num="APCS", grade=score, creds=4, desc=course_desc, ap=True, ap_ppf_desc="CS AP")
    # AP Calc BC
    elif ap_desc == "Mathematics: Calculus BC" and (score == "4" or score == "5"):
        equiv = Course(num="APCALC", grade=score, creds=4, desc=course_desc, ap=True, ap_ppf_desc="MATH AP")
    # AP Physics
    elif ap_desc == "Physics C - Mechanics" and score == "5":
        equiv = Course(num="APMECH", grade=score, creds=4, desc=course_desc, ap=True, ap_ppf_desc="PHYS AP")
    elif ap_desc == "Physics C - Electricity & Magt" and score == "5":
        equiv = Course(num="APELECTRO", grade=score, creds=4, desc=course_desc, ap=True, ap_ppf_desc="PHYS AP")

    # AP English
    elif ap_desc == "English Language &Composition" and (score == "4" or score == "5"):
        equiv = Course(num="APENGLANG", grade=score, creds=3, desc=course_desc, ap=True, ap_ppf_desc="ENGL AP")
    elif ap_desc == "English Literature &Compostn" and (score == "4" or score == "5"):
        equiv = Course(num="APENGLIT", grade=score, creds=3, desc=course_desc, ap=True, ap_ppf_desc="ENGL AP")

    # AP Languages
    elif ap_desc == "Spanish Language" and (score == "4" or score == "5"):
        equiv = Course(num="APSPANLANG", grade=score, creds=3, desc=course_desc, ap=True, ap_ppf_desc="SPAN AP")
    elif ap_desc == "Spanish Literature" and (score == "4" or score == "5"):
        equiv = Course(num="APSPANLIT", grade=score, creds=3, desc=course_desc, ap=True, ap_ppf_desc="SPAN AP")
    elif ap_desc == "French Language" and (score == "4" or score == "5"):
        equiv = Course(num="APFRENLANG", grade=score, creds=3, desc=course_desc, ap=True, ap_ppf_desc="FREN AP")
    elif ap_desc == "French Literature" and (score == "4" or score == "5"):
        equiv = Course(num="APFRENLIT", grade=score, creds=3, desc=course_desc, ap=True, ap_ppf_desc="FREN AP")
    elif ap_desc == "German Language & Culture" and (score == "4" or score == "5"):
        equiv = Course(num="APGERMAN", grade=score, creds=3, desc=course_desc, ap=True, ap_ppf_desc="GERM AP")
    elif ap_desc == "Italian Language & Culture" and (score == "4" or score == "5"):
        equiv = Course(num="APITALIAN", grade=score, creds=3, desc=course_desc, ap=True, ap_ppf_desc="ITAL AP")

    # Economics & Psychology
    elif ap_desc == "Economics: Macroeconomics" and (score == "4" or score == "5"):
        equiv = Course(num="APMACRO", grade=score, creds=3, desc=course_desc, ap=True, ap_ppf_desc="ECON AP")
    elif ap_desc == "Economics: Microeconomics" and (score == "4" or score == "5"):
        equiv = Course(num="APMICRO", grade=score, creds=3, desc=course_desc, ap=True, ap_ppf_desc="ECON AP")
    elif ap_desc == "Psychology" and (score == "4" or score == "5"):
        equiv = Course(num="APPSYCH", grade=score, creds=3, desc=course_desc, ap=True, ap_ppf_desc="PSYCH AP")

    else:
        return None

    return equiv


def read_class_from_transcript(transcript, cols, row):

    ap_loc = cols['ap y/n'] + row
    transfer_loc = cols['transfer'] + row

    if transcript[transfer_loc].value is not None:
        return transcript[transfer_loc].value

    if transcript[ap_loc].value is None:

        dept_loc = cols['dept'] + row
        dept = transcript[dept_loc].value

        num_loc = cols['num'] + row
        num = transcript[num_loc].value

        course_num = dept + str(num)

        creds_loc = cols['creds'] + row
        creds = transcript[creds_loc].value

        grade_loc = cols['grade'] + row
        grade = transcript[grade_loc].value
        if grade is not None and "*" in grade:
            grade = grade[:-1]

        desc_loc = cols['desc'] + row
        desc = transcript[desc_loc].value

        term_loc = cols['semester'] + row
        term = transcript[term_loc].value

        c = Course(course_num, grade, creds, term=term, desc=desc)

    else:

        ap_desc_loc = cols['ap desc'] + row
        ap_desc = transcript[ap_desc_loc].value

        score_loc = cols['ap score'] + row
        score = str(transcript[score_loc].value)

        course_desc_loc = cols['desc'] + row
        course_desc = transcript[course_desc_loc].value

        c = determine_ap_course_equiv(ap_desc, score, course_desc)

    return c


def upload_courses_from_file(filename):
    courses = []
    wb = xlrd.open_workbook(filename)
    sheet = wb.sheet_by_index(0)
    for i in range(1, sheet.nrows):
        courses.append(sheet.cell_value(i, 0))
    return courses


def is_semester(s):
    if s is None:
        return True
    elif isinstance(s, str) is False:
        return False
    else:
        return any(c.isdigit() for c in s)


def categories_represented(courses):
    total_cats = []
    for c in courses:
        for cat in c.categories:
            if total_cats.__contains__(cat) is False:
                total_cats.append(cat)
    table = np.zeros([len(courses), len(total_cats)], dtype=bool)
    for c in courses:
        for l in total_cats:
            if c.categories.__contains__(l):
                table[courses.index(c)][total_cats.index(l)] = 1
    tot = 0
    for c in range(len(courses)):
        if sum(table[c][:]) >= 1:
            tot += 1
    return tot


def sdescr2ldescr(s):
    assert " " in s is False, "Is this really a short description? %s\n" % s
    year = s[0:4]
    sem = s[4:]

    if sem == "FA":
        return "Fall " + year
    elif sem == "SP":
        return "Spring " + year
    else:
        print("%s is not an expected sdescr (XXXXFA or XXXXSP)" % s)


def ldescr2sdescr(s):
    assert " " in s, "Is this really a long description? %s\n" % s
    [sem, year] = s.split()

    if sem == "Fall":
        return year + "FA"
    elif sem == "Spring":
        return year + "SP"
    else:
        print("%s is not an expected ldescr (Fall XXXX or Spring XXXX)" % s)